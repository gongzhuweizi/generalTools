# -*-coding:utf-8 -*-
"""
@Author: zhangyi
@Time: 2021/7/12 10:55 上午
@File: OperateExcel.py
@IDE: PyCharm

1. 字典按照列进行导入默认不添加表头，可以通过setExcelHeaders函数进行设置，传入类型为list。列表按照行导入无需设置表头，指定columns即可
2. 默认生成、追加、修改的excel路径为当前目录，文件文成为demo-2021-08-01.xls，2021-08-01为当前日。可以通过setExcelDirFileName进行配置。
3. 默认为sheet1，可以通过setSheetName函数进行配置
4. excel默认索引为False，如果需要配置请调用setIndex函数
"""
import pandas as pd
import Tools
import openpyxl
from openpyxl import Workbook,load_workbook
from Tools import fileExsits

"""
操作excel工具类
"""
class OperateExcel():
	excelHeaders = True #默认添加表头
	excelDirFileName = 'excelData/demo-'+Tools.getCurrentTimeYMD()+".xlsx" #例如：默认保存路径为当前目录下的demo-2021-08-05.xlsx
	sheetName = 'Sheet1' #默认sheet1中保存
	excelIndex = False #默认不添加索引
	excelColumns = None
	writer = None



	"""
	设置excel文件路径已经文件名
	"""
	def setExcelDirFileName(self,excelDirFileName)->None:
		self.excelDirFileName = excelDirFileName

	"""
	设置excel保存的数据，参数为dict格式
	"""
	def setExcelData(self,excelDate)->None:
		self.excelData = excelDate

	"""
	设置保存excel的sheet名称，默认为sheet1
	"""
	def setSheetName(self,sheetName)->None:
		self.sheetName = sheetName
	"""
	设置excel表头，可以不设置，默认为没有表头，适用于字典按照列进行配置list类型
	"""
	def setExcelHeaders(self,excelHeaders)->None:
		if isinstance(excelHeaders,bool):
			self.excelHeaders = excelHeaders
		else:
			print("ERROR!传入的标题字段不是true或者false类型")
	def setExcelColumns(self,excelColumns)->None:
			self.excelColumns  = excelColumns

	"""
	设置excel索引，可以不设置，默认没有索引，list类型
	"""
	def setIndex(self,excelIndex)->None:
		self.excelIndex = excelIndex

	"""
	获取pd加载的数据对象
	"""
	def __getExcelObject(self) -> pd.DataFrame:
		df = pd.DataFrame(self.excelData,
						  columns = self.excelColumns
						  )
		return df

	"""
	pandas写入excel数据，在写入excel指定sheet名称、是否需要index
	"""
	def __writeExcel(self,df)->None:
		df.to_excel(self.excelDirFileName,
					header = self.excelHeaders,
					sheet_name=self.sheetName,
					index = self.excelIndex,
					excel_writer = self.writer
					)
	def __createWriteExcel(self)->pd.DataFrame:
		df = self.__getExcelObject()
		return df

	"""
	创建excel的时候调用，按照列进行写入,单次，不可追加
	"""
	def pdColumnDictWrite(self)->None:
		if isinstance(self.excelData,dict):
			df  = self.__createWriteExcel()
			self.__writeExcel(df)

		else:
			print("ERROR!传入excel数据不是dict类型")
	"""
	创建excel的时候调用，按照行进行写入，单次，不可追加,data必须是双list类型
	"""
	def pdLineListWrite(self)->None:
		if isinstance(self.excelData[0],list):
			df = self.__createWriteExcel()
			self.__writeExcel(df)
		else:
			print("ERROR!传入excel数据不是二维list类型")
	"""
	读取excel
	"""
	def readExcelFileName(self):
		df = pd.read_excel(self.excelDirFileName)
		return df

	"""
	追加写入，按照行进行写入，可以追加，默认为
	"""
	def pdAppendListWrite(self):
		ds = self.__getExcelObject()
		df = self.readExcelFileName()
		print(df)
		d = df.append(ds,ignore_index=True)
		self.__writeExcel(d)
	"""
	增加sheet，并写入数据
	"""
	def addExcelSheetWrite(self):
		if fileExsits(self.excelDirFileName):
			wb = Workbook()
			sh = wb.create_sheet(self.sheetName)
		wb = load_workbook(self.excelDirFileName)













